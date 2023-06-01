from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import load_workbook


def convert_xlsx_to_pdf(xlsx_file, pdf_file):
    # Load the XLSX file
    workbook = load_workbook(filename=xlsx_file)
    worksheet = workbook.active

    # Create a PDF canvas
    pdf_canvas = canvas.Canvas(pdf_file, pagesize=letter)

    # Set the font and font size
    pdf_canvas.setFont("Helvetica", 12)

    # Iterate through each merged cell range in the worksheet
    for merged_range in worksheet.merged_cells.ranges:
        # Get the top-left cell coordinates of the merged range
        start_cell = merged_range.min_row, merged_range.min_col

        # Get the value from the top-left cell of the merged range
        value = worksheet.cell(*start_cell).value

        # Convert None values to empty strings
        if value is None:
            value = ""

        # Calculate the PDF canvas coordinates based on the top-left cell coordinates
        x = (start_cell[1] - 1) * 60
        y = letter[1] - (start_cell[0] * 20)

        # Write the cell value to the PDF canvas
        pdf_canvas.drawString(x, y, str(value))

    # Save the PDF file
    pdf_canvas.save()


# Example usage
convert_xlsx_to_pdf(r"F:\Code\Python-Projects\Laser-Quote-Generator\quotes\2023-05-31-17-55-49.xlsx", "output.pdf")
