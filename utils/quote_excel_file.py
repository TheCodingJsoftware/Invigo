import contextlib
import re
from datetime import datetime

import xlsxwriter
from openpyxl.utils.cell import column_index_from_string, get_column_letter


class ExcelFile:
    def __init__(self, file_name: str, generate_quote: bool, should_generate_packing_slip: bool) -> None:
        self.workbook = xlsxwriter.Workbook(file_name)
        self.workbook.set_properties(
            {
                "title": "Laser Quote",
                "subject": "Quote for parts",
                "author": "Jared Gross",
                "manager": "Jared Gross",
                "company": "TheCodingJ'software",
                "category": "Laser Quotes",
                "keywords": "Laser, Quotes, Laser Quotes",
                "comments": "Created with Python, Magic, XlsxWriter and most importantly, Love",
            }
        )
        self.FONT_NAME: str = "Book Antiqua"
        self.worksheet = self.workbook.add_worksheet("Sheet")
        self.info_worksheet = self.workbook.add_worksheet("info")
        self.worksheet.hide_gridlines(2)
        self.worksheet.set_margins(0.25, 0.25, 0.25, 0.55)
        footer = "&RPage &P of &N"
        self.worksheet.set_footer(footer)
        self.generate_quote: bool = generate_quote
        self.should_generate_packing_slip: bool = should_generate_packing_slip

        self.cell_regex = r"^([A-Z]+)([1-9]\d*)$"
        self.file_name = file_name

    def parse_cell(self, cell: str):
        """Parses excel cell input such as "AD300"

        Args:
            cell (str): input -> "AD300"

        Returns:
            str: "AD"
            int: 300
        """
        cell = cell.upper()
        if matches := re.search(self.cell_regex, cell):
            return (matches[1], int(matches[2]))

    def add_list_to_sheet(self, cell: str, items: list, horizontal: bool = True) -> None:
        """Adds a list of items to the specfied sheet
        Args:
            sheet_name (str): Name of the sheet you want to add a list to.
            cell (str): specfied cell location, such as "A1"
            items (list): any list of items you want to add to the excel sheet
            horizontal (bool, optional): Allows for inputing lists vertical(False) or horizontal(True). Defaults to True.
        """
        col, row = self.parse_cell(cell=cell)
        if horizontal:
            col_index = column_index_from_string(col)
            for item in items:
                col_str = get_column_letter(col_index)
                try:
                    if item.is_integer():
                        self.add_item_to_sheet(f"{col_str}{row}", item)
                    elif not item.is_integer():
                        self.add_item_to_sheet(f"{col_str}{row}", item)
                except AttributeError:
                    self.add_item_to_sheet(f"{col_str}{row}", item)
                col_index += 1
        else:
            for item in items:
                try:
                    if item.is_integer():
                        self.add_item_to_sheet(f"{col}{row}", item)
                    elif not item.is_integer():
                        self.add_item_to_sheet(f"{col}{row}", item)
                except AttributeError:
                    self.add_item_to_sheet(f"{col}{row}", item)
                row += 1

    def add_item_to_sheet(self, cell: str, item, number_format=None) -> None:
        """Add any item to any cell in the specified sheet
        Args:
            cell (str): Such as "A1"
            item (any): Any (item, str, int, float)
        """
        col, row = self.parse_cell(cell=cell)

        cell_format = self.workbook.add_format()
        self.info_worksheet.set_column("J:J", 12)
        self.info_worksheet.set_column("L:L", 12)
        with contextlib.suppress(Exception):
            if "NOW" in item:
                cell_format = self.workbook.add_format({"num_format": "hh:mm:ss AM/PM"})
        try:
            if item.is_integer():
                self.info_worksheet.write(f"{col}{row}", int(item), cell_format)
            elif not item.is_integer():
                self.info_worksheet.write(f"{col}{row}", float(item), cell_format)
        except AttributeError:
            self.info_worksheet.write(f"{col}{row}", item, cell_format)

    def set_row_hidden_sheet(self, cell: str, hidden: bool = True) -> None:
        """Hide row

        Args:
            cell (str): Such as "A1"
            visible (bool): True or False
        """
        _, row = self.parse_cell(cell=cell)
        self.info_worksheet.set_row(row - 1, None, None, {"hidden": 1})

    def add_list(self, cell: str, items: list, horizontal: bool = True) -> None:
        """Adds a list of items to the current workbook

        Args:
            sheet_name (str): Name of the sheet you want to add a list to.
            cell (str): specfied cell location, such as "A1"
            items (list): any list of items you want to add to the excel sheet
            horizontal (bool, optional): Allows for inputing lists vertical(False) or horizontal(True). Defaults to True.
        """
        col, row = self.parse_cell(cell=cell)
        if horizontal:
            col_index = column_index_from_string(col)
            for item in items:
                col_str = get_column_letter(col_index)
                self.add_item(f"{col_str}{row}", item)
                col_index += 1
        else:
            for item in items:
                self.add_item(f"{col}{row}", item)
                row += 1

    def add_item(self, cell: str, item, number_format=None, totals: bool = False) -> None:
        """Add any item to any cell in the excel work book

        Args:
            cell (str): Such as "A1"
            item (any): Any (item, str, int, float)
        """
        col, row = self.parse_cell(cell=cell)

        if number_format is None:
            cell_format = self.workbook.add_format({"font_name": self.FONT_NAME})
        else:
            cell_format = self.workbook.add_format({"num_format": number_format, "font_name": self.FONT_NAME})
        if "Payment" not in str(item) and "Received" not in str(item) and "__" not in str(item):
            if "Sheet Count:" not in str(item):
                cell_format.set_align("center")
                cell_format.set_align("vcenter")
                cell_format.set_text_wrap()
        if (
            "Total" in str(item)
            or "Packing Slip" in str(item)
            or "Order #" in str(item)
            or "Ship To:" in str(item)
            or "Date Shipped:" in str(item)
            or "No Tax Included" in str(item)
            or "=SUM(Table1[Price])" in str(item)
            or "TEXTAFTER" in str(item)
        ):
            cell_format.set_bold()
        if col == "K" and row > 2 and "Tax" not in str(item):
            cell_format.set_right(1)
        if col == "G" and not self.generate_quote and row > 4:
            cell_format.set_right(1)
        if totals:
            cell_format.set_top(6)
            cell_format.set_bottom(1)
            if col == "A":
                cell_format.set_left(1)
        try:
            if item.is_integer():
                self.worksheet.write(f"{col}{row}", int(item), cell_format)
            elif not item.is_integer():
                self.worksheet.write(f"{col}{row}", float(item), cell_format)
        except (TypeError, AttributeError):
            self.worksheet.write(f"{col}{row}", item, cell_format)

    def set_cell_width(self, cell: str, width: int) -> None:
        # sourcery skip: remove-unnecessary-cast
        """Change teh width of any cell, only the column is, the row is not used.

        Args:
            cell (str): Such as "A1"
            width (int): The width you want that column to be
        """
        col, _ = self.parse_cell(cell=cell)
        self.worksheet.set_column(f"{col}:{col}", int(width))

    def set_cell_height(self, cell: str, height: int) -> None:
        """Change teh width of any cell, only the row is, the column is not used.

        Args:
            cell (str): Such as "A1"
            height (int): The height you want that row to be
        """
        _, row = self.parse_cell(cell=cell)

        self.worksheet.set_row(row - 1, height)

    def add_image(self, cell: str, path_to_image: str) -> None:
        """Add an image to any cell

        Args:
            cell (str): Such as "A1"
            path_to_image (str): The direct path to the image
        """
        col, row = self.parse_cell(cell=cell)
        self.worksheet.insert_image(
            f"{col}{row}",
            path_to_image,
            {"x_offset": 2, "y_offset": 2, "x_scale": 1, "y_scale": 1},
        )

    def add_dropdown_selection(self, cell: str, type: str, location: str) -> None:
        """Add a data validation drop down selection for any cell

        Args:
            cell (str): Such as "A1"
            type (str): 'list'
            formula (str): the location of where the list is located such as: "A1:C1"
        """
        col, row = self.parse_cell(cell=cell)
        self.worksheet.data_validation(f"${col}${row}", {"validate": type, "source": location})

    def add_table(self, display_name: str, theme: str, location: str, headers: list) -> None:
        """Add a table to the excel sheet

        Args:
            display_name (str): Name of that table, such as "Table1"
            theme (str): Any color theme provided by excel itself
            location (str): The location you want to format the table, such as: "A1:B3"
        """
        self.worksheet.add_table(
            location,
            {
                "total_row": True,
                "style": theme,
                "first_column": False,
                "columns": [{"header": header} for header in headers],
            },
        )

    def set_pagebreak(self, row: int) -> None:
        """
        This function sets a horizontal page break at a specified row in a worksheet.

        Args:
          row (int): The "row" parameter is an integer that represents the row number where the page
        break should be set in the worksheet. The method "set_h_pagebreaks" is used to set horizontal
        page breaks in the worksheet, and it takes a list of row numbers as its argument. In this case,
        """
        self.worksheet.set_h_pagebreaks([row])

    def set_col_hidden(self, cell: str, hidden: bool = True) -> None:
        """Hide column

        Args:
            cell (str): Such as "A1"
            visible (bool): True or False
        """
        col, _ = self.parse_cell(cell=cell)
        self.worksheet.set_column(f"{col}:{col}", None, None, {"hidden": 1})

    def set_row_hidden(self, cell: str, hidden: bool = True) -> None:
        """Hide row

        Args:
            cell (str): Such as "A1"
            visible (bool): True or False
        """
        _, row = self.parse_cell(cell=cell)
        self.worksheet.set_row(row, None, None, {"hidden": 1})

    def add_macro(self, macro_path) -> None:
        self.workbook.add_vba_project(macro_path)

    def set_print_area(self, cell) -> None:
        self.worksheet.print_area(cell)

    def freeze_pane(self, row: int) -> None:
        self.worksheet.freeze_panes(f"A{row}")

    def save(self) -> None:
        """Save excel file."""
        merge_format = self.workbook.add_format({"align": "top", "valign": "right", "font_name": self.FONT_NAME})
        merge_format.set_text_wrap()
        self.worksheet.merge_range("J1:K1", f"{datetime.now().strftime('%B %d, %A, %Y')}", merge_format)
        merge_format = self.workbook.add_format({"align": "center", "valign": "center", "font_name": self.FONT_NAME})
        merge_format.set_bold()
        merge_format.set_font_size(18)
        merge_format.set_bottom(1)

        if not self.generate_quote:
            self.worksheet.merge_range("E1:G1", "Work Order", merge_format)
        else:
            self.worksheet.merge_range("E1:G1", "Quote", merge_format)
        if self.should_generate_packing_slip:
            self.worksheet.merge_range("E1:G1", "Packing Slip", merge_format)
        self.workbook.close()
