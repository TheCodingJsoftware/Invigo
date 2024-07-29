from openpyxl import Workbook, load_workbook


class PriceHistoryFile:
    def __init__(self, file_name: str):
        self.file_name = file_name
        self.row_pos: int = 0
        try:
            self.workbook = load_workbook(f"Price History Files/{self.file_name}")
            self.price_history_sheet = self.workbook["Price History"]
        except Exception:
            self.workbook = Workbook()
            self.workbook.create_sheet("Price History", 0)
            self.workbook.remove_sheet(self.workbook["Sheet"])
            self.price_history_sheet = self.workbook["Price History"]
            self.add_new("Date", "Part Name", "Part #", "Old Price", "New Price")
        self.data = {
            "Date": [],
            "Part Name": [],
            "Part Number": [],
            "Old Price": [],
            "New Price": [],
        }
        self.load_file()

    def load_file(self):
        self.row_pos = len(self.price_history_sheet["A"])

        for cell in self.price_history_sheet["A"]:
            self.data["Date"].append(cell.value)
        for cell in self.price_history_sheet["B"]:
            self.data["Part Name"].append(cell.value)
        for cell in self.price_history_sheet["C"]:
            self.data["Part Number"].append(cell.value)
        for cell in self.price_history_sheet["D"]:
            self.data["Old Price"].append(cell.value)
        for cell in self.price_history_sheet["E"]:
            self.data["New Price"].append(cell.value)

    def get_data_from_category(self) -> dict:
        return self.data

    def add_new(
        self,
        date: str,
        part_name: str,
        part_number: str,
        old_price: float,
        new_price: float,
    ):
        if old_price == new_price:
            return
        self.price_history_sheet.cell(row=self.row_pos + 1, column=1, value=date)
        self.price_history_sheet.cell(row=self.row_pos + 1, column=2, value=part_name)
        self.price_history_sheet.cell(row=self.row_pos + 1, column=3, value=part_number)
        self.price_history_sheet.cell(row=self.row_pos + 1, column=4, value=old_price)
        self.price_history_sheet.cell(row=self.row_pos + 1, column=5, value=new_price)
        self.workbook.save(f"Price History Files/{self.file_name}")
