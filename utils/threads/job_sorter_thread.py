import os
import re
import shutil
import traceback
from itertools import islice

from openpyxl import load_workbook
from PIL import Image
from PyQt6.QtCore import QThread, pyqtSignal


class JobSorterThread(QThread):
    signal = pyqtSignal(object)

    def __init__(
        self,
        parent,
        job_name: str,
        path_to_excel_file: str,
        directory_to_sort: str,
        output_directory: str,
    ) -> None:
        QThread.__init__(self)
        self.data = {}
        self.path_to_excel_file: str = path_to_excel_file
        self.directory_to_sort: str = directory_to_sort
        self.job_name: str = job_name
        self.output_directory: str = os.path.join(output_directory, self.job_name)

    def _make_name_safe(self, name: str):
        invalid_chars_pattern = r'[<>:"/\\|?*]'

        return re.sub(invalid_chars_pattern, "_", name)

    def _find_column(self, path_to_excel_file: str, column_name_to_find: str) -> int:
        workbook = load_workbook(path_to_excel_file)
        worksheet = workbook.active

        try:
            for column, cell in enumerate(worksheet[1]):
                if name := cell.value:
                    if column_name_to_find.lower() in name.lower():
                        return column
        except TypeError:
            return -1
        return -1

    def get_data_from_excel(self, path_to_excel_file: str) -> dict:
        workbook = load_workbook(filename=path_to_excel_file)
        worksheet = workbook.active
        part_names_new_with_thickness = {}
        part_name_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Part")
        material_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Material")
        thickness_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Thick")
        quantity_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Parts Per")
        if quantity_column == -1:
            quantity_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Qty")
        for row in islice(worksheet.iter_rows(values_only=True), 1, None):
            part_name: str = row[part_name_column]
            material: str = row[material_column]
            thickness: str = row[thickness_column]
            quantity: str = row[quantity_column]

            part_names_new_with_thickness[part_name] = {
                "thickness": self._make_name_safe(str(thickness)),
                "quantity": quantity,
            }

        return part_names_new_with_thickness

    def get_all_file_paths_from_directory(self, directory_to_sort: str) -> list[str]:
        file_paths = []
        for root, directories, files in os.walk(directory_to_sort):
            for file in files:
                file_path = os.path.join(root, file)
                file_paths.append(file_path)
        return file_paths

    def filter_file_paths(self, file_paths: list[str], desired_file_names: list[str]) -> list[str]:
        filtered_paths = []
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            name_without_extension = os.path.splitext(file_name)[0]
            if name_without_extension in desired_file_names:
                filtered_paths.append(file_path)
        return filtered_paths

    def run(self) -> None:
        try:
            data = self.get_data_from_excel(self.path_to_excel_file)
            all_file_paths = self.get_all_file_paths_from_directory(directory_to_sort=self.directory_to_sort)
            filtered_paths = self.filter_file_paths(all_file_paths, list(data.keys()))
            new_copy_location = {}
            for file_path in filtered_paths:
                file_name: str = os.path.basename(file_path)
                name_without_extension: str = os.path.splitext(file_name)[0]
                extension: str = os.path.splitext(file_name)[1].replace(".", "").upper()
                quantity_mulitplier: str = f" x{data[name_without_extension]['quantity']}" if int(data[name_without_extension]["quantity"]) > 1 else ""
                new_copy_location[file_path] = {
                    "new_location": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}",
                    "old_name": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{file_name}",
                    "new_name": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{name_without_extension}{quantity_mulitplier}{os.path.splitext(file_name)[1]}",
                }

            for original_location in new_copy_location:
                new_location = new_copy_location.get(original_location)["new_location"]
                old_name = new_copy_location.get(original_location)["old_name"]
                new_name = new_copy_location.get(original_location)["new_name"]
                os.makedirs(new_location, exist_ok=True)
                shutil.copy(original_location, new_location)
                if "DXF" in new_location:
                    try:
                        os.rename(old_name, new_name)
                    except FileExistsError:
                        os.remove(old_name)

            self.signal.emit("Done")
        except Exception as e:
            self.signal.emit(f"ERROR!\nException: {e}\nTrace stack:\n{traceback.print_exc()}\n\n")
