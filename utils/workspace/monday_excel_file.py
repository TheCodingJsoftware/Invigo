from itertools import zip_longest

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from utils.workspace.workspace_item import WorkspaceItem


class MondayExcelFile:
    def __init__(self, path: str):
        self.path: str = path
        self.columns_to_find: list[str] = [
            "name",
            "paint color",
            "thickness",
            "material type",
            "parts per",
        ]  # in lower
        self.data: dict[str, dict[dict[str, int], dict[str, any]]] = {}

    def find_jobs(self, sheet: Worksheet):
        for row_index in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=1)
            cell_value = cell.value
            if not cell_value or not isinstance(cell_value, str):
                continue
            if cell_value.lower() in self.columns_to_find and cell_value.lower() == "name":
                cell_above = sheet.cell(row=cell.row - 1, column=cell.column)  # Job name
                self.data[cell_above.value] = {
                    "job_data": {"start_row": cell.row},
                    "items": {},
                }

    def find_table_headers(self, sheet: Worksheet):
        for job_index, (job_name, job_data) in enumerate(self.data.items()):
            job_row: int = job_data["job_data"]["start_row"]
            try:
                end_row = self.data[list(self.data.keys())[job_index + 1]]["job_data"]["start_row"] - 3  # needs to offset to the next job name
            except IndexError:
                end_row = sheet.max_row
            for col_index in range(1, sheet.max_column):
                header_cell = sheet.cell(row=job_row, column=col_index)
                if header_cell.value.lower() in self.columns_to_find:
                    job_data["items"][header_cell.value.lower()] = []
                    for row in range(job_row + 1, end_row):
                        cell_item = sheet.cell(row=row, column=col_index)
                        job_data["items"][header_cell.value.lower()].append(cell_item.value)

    def clean_up_data(self) -> dict[str, dict[str, any]]:
        data: dict[str, dict[str, any]] = {}
        for job_name, job_data in self.data.items():
            data[job_name] = {}
            # for headers, item_data in job_data["items"].items():
            for (
                item_name,
                paint_color,
                thickness,
                material_type,
                parts_per,
            ) in zip_longest(
                job_data["items"]["name"],
                job_data["items"]["paint color"],
                job_data["items"]["thickness"],
                job_data["items"]["material type"],
                job_data["items"]["parts per"],
                fillvalue=None,
            ):
                if item_name is None:
                    continue
                # Its a formula, it should be float or int
                if isinstance(parts_per, str) or item_name == "Subitems" or (thickness == material_type == None):
                    continue
                data[job_name][item_name] = {
                    "paint_color": paint_color,
                    "thickness": thickness,
                    "material_type": material_type,
                    "parts_per": parts_per,
                }
        return self.remove_empty_jobs(data)

    def remove_empty_jobs(self, data: dict[str, dict[str, any]]) -> dict[str, dict[str, any]]:
        new_data: dict[str, dict[str, any]] = {job_name: job_data for job_name, job_data in data.items() if job_data}
        return new_data

    def combine_data(self, data: dict[str, dict[str, any]]) -> dict[str, list[WorkspaceItem]]:
        new_data: dict[str, list[WorkspaceItem]] = {}
        for job_name, job_data in data.items():
            new_data[job_name] = []
            for item_name, item_data in job_data.items():
                item: WorkspaceItem = WorkspaceItem(
                    name=item_name,
                    data={
                        "Bending Files": [],
                        "Welding Files": [],
                        "CNC/Milling Files": [],
                        "thickness": item_data["thickness"],
                        "material": item_data["material_type"],
                        "paint_type": None,
                        "paint_color": item_data["paint_color"],
                        "parts_per": item_data["parts_per"],
                        "flow_tag": [],
                        "timers": {},
                        "customer": "",
                        "ship_to": "",
                        "show": True,
                        "notes": "",
                    },
                )
                new_data[job_name].append(item)
        return new_data

    def get_data(self) -> dict[str, list[WorkspaceItem]]:
        workbook = openpyxl.load_workbook(filename=self.path)
        sheet = workbook.active
        self.find_jobs(sheet=sheet)
        self.find_table_headers(sheet=sheet)
        return self.combine_data(self.clean_up_data())
