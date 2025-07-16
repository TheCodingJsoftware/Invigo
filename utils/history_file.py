from openpyxl import Workbook, load_workbook


class HistoryFile:
    def __init__(self):
        self.file_name = "inventory history.xlsx"
        self.category_new_row_pos: int = 0
        self.single_item_new_row: int = 0
        try:
            self.workbook = load_workbook(f"data/{self.file_name}")
        except Exception:
            self.workbook = Workbook()
            self.workbook.create_sheet("Categories", 0)
            self.workbook.create_sheet("Single Items", 1)
            self.workbook.remove_sheet(self.workbook["Sheet"])
        self.categories_sheet = self.workbook["Categories"]
        self.single_items_sheet = self.workbook["Single Items"]
        self.category_data = {"Date": [], "Description": []}
        self.single_item_data = {"Date": [], "Description": []}
        self.load_file()

    def load_file(self):
        self.category_new_row_pos = len(self.categories_sheet["A"])

        for cell in self.categories_sheet["A"]:
            self.category_data["Date"].append(cell.value)
        for cell in self.categories_sheet["B"]:
            self.category_data["Description"].append(cell.value)

        self.single_item_new_row = len(self.single_items_sheet["A"])

        for cell in self.single_items_sheet["A"]:
            self.single_item_data["Date"].append(cell.value)
        for cell in self.single_items_sheet["B"]:
            self.single_item_data["Description"].append(cell.value)

    def get_data_from_category(self) -> dict:
        return self.category_data

    def get_data_from_single_item(self) -> dict:
        return self.single_item_data

    def add_new_to_category(self, date: str, description: str):
        self.categories_sheet.cell(row=self.category_new_row_pos + 1, column=1, value=date)
        self.categories_sheet.cell(row=self.category_new_row_pos + 1, column=2, value=description)
        self.workbook.save(f"data/{self.file_name}")
        self.load_file()

    def add_new_to_single_item(self, date: str, description: str):
        self.single_items_sheet.cell(row=self.single_item_new_row + 1, column=1, value=date)
        self.single_items_sheet.cell(row=self.single_item_new_row + 1, column=2, value=description)
        self.workbook.save(f"data/{self.file_name}")
        self.load_file()
