import os
import shutil
import time
from datetime import date

from openpyxl import load_workbook


class POTemplate:
    def __init__(self, po_template: str):
        self.po_template = po_template
        self.order_number_cell = (4, 6)  # F4
        self.date_cell = (6, 6)  # F6
        self.cwd: str = os.path.abspath(os.getcwd()).replace("\\", "/")
        self.date = date.today().strftime("%m/%d/%y")
        self.signature: str = f"Lynden Gross                                                          {self.date}"
        self.vendor: str = self.get_vendor()
        self.order_number: int = self.get_order_number()

    def copy_file_with_overwrite(self, source, target, retry_interval=1, max_retries=10):
        retries = 0
        while retries < max_retries:
            try:
                if os.path.exists(target):
                    if os.path.samefile(source, target):
                        os.remove(target)
                shutil.copyfile(source, target)
                return
            except shutil.SameFileError:
                if os.path.samefile(source, target):
                    os.remove(target)
                    shutil.copyfile(source, target)
                    return
            except PermissionError as e:
                if e.winerror == 32:  # File in use error
                    retries += 1
                    time.sleep(retry_interval)
                else:
                    raise
        raise PermissionError(f"Failed to copy file {source} to {target} after {max_retries} retries.")

    def generate(self):
        self.output_path: str = f"{self.cwd}/PO's/{self.vendor}/PO {self.order_number+1}.xlsx"
        self.copy_file_with_overwrite(self.po_template, self.output_path)
        self.set_order_number()
        self.set_date()
        self.set_signature()

    def get_vendor(self) -> str:
        workbook = load_workbook(filename=self.po_template)
        worksheet = workbook.active

        columns_to_search: list[str] = ["B", "C"]
        vendor_col: int = 0
        for column_number, column in enumerate(columns_to_search, start=1):
            for col in worksheet[column]:
                if col.value is None:
                    continue
                if "Vendor" in str(col.value):
                    vendor_col = column_number + 1  # They dont start at 0
        return " ".join((str(worksheet.cell(row=row + 2, column=vendor_col).value).replace("\n", " ").replace("None", "").replace("  ", " ")) for row in range(4)).strip()

    def get_order_number(self) -> int:
        order_number: int = None
        workbook = load_workbook(filename=self.po_template)
        worksheet = workbook.active
        order_number = int(worksheet.cell(row=self.order_number_cell[0], column=self.order_number_cell[1]).value)  # Merged: F4:G4
        # We only want to update the master templates PO number if it
        # was loaded from the templates directory, otherwise this is a new
        # file being added
        if "PO's" in self.po_template:
            worksheet.cell(row=self.order_number_cell[0], column=self.order_number_cell[1]).value = order_number + 1
            workbook.save(self.po_template)
            workbook.close()
        return order_number

    def get_output_path(self) -> str:
        return self.output_path

    def set_order_number(self):
        workbook = load_workbook(filename=self.output_path)
        worksheet = workbook.active
        worksheet.cell(row=self.order_number_cell[0], column=self.order_number_cell[1]).value = self.order_number + 1
        workbook.save(self.output_path)
        workbook.close()

    def set_date(self):
        workbook = load_workbook(filename=self.output_path)
        worksheet = workbook.active
        worksheet.cell(row=self.date_cell[0], column=self.date_cell[1]).value = self.date
        workbook.save(self.output_path)
        workbook.close()

    def set_signature(self):
        workbook = load_workbook(filename=self.output_path)
        worksheet = workbook.active
        signature_row: int = 0
        for cell in worksheet["E"]:
            if cell.value is None:
                continue
            if "Authorized by" in str(cell.value):
                signature_row = cell.row - 1
        worksheet.cell(row=signature_row, column=5).value = self.signature  # E:{signature_row}
        workbook.save(self.output_path)
        workbook.close()


if __name__ == "__main__":
    p = POTemplate(r"F:\Code\Python-Projects\Inventory Manager\Piney MFG\Cloverdale Paint\Clover Dale Paints PO#.xlsx")
    print(p.get_vendor())
