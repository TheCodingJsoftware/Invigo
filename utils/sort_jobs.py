import os
import re
import shutil
from itertools import islice

import PyPDF2
import tabula
from openpyxl import load_workbook


def _make_name_safe(name: str):
    invalid_chars_pattern = r'[<>:"/\\|?*]'
    return re.sub(invalid_chars_pattern, "_", name)


def _find_column(path_to_excel_file: str, column_name_to_find: str) -> int:
    workbook = load_workbook(path_to_excel_file)
    worksheet = workbook.active

    try:
        for column, cell in enumerate(worksheet[1]):
            if column_name_to_find in cell.value:
                return column
    except TypeError:
        return -1
    return -1


def get_data_from_excel(path_to_excel_file: str) -> dict:
    workbook = load_workbook(filename=path_to_excel_file)
    worksheet = workbook.active
    part_names_new_with_thickness = {}
    part_name_column: int = _find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Part")
    material_column: int = _find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Material")
    thickness_column: int = _find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Thick")
    quantity_column: int = _find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Parts Per")
    if quantity_column == -1:
        quantity_column: int = _find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Qty")
    for row in islice(worksheet.iter_rows(values_only=True), 1, None):
        part_name: str = row[part_name_column]
        material: str = row[material_column]
        thickness: str = row[thickness_column]
        quantity: str = row[quantity_column]

        part_names_new_with_thickness[part_name] = {
            "thickness": _make_name_safe(str(thickness)),
            "quantity": quantity,
        }

    return part_names_new_with_thickness


def get_data_from_pdf(path_to_pdf_file: str) -> dict:
    part_names_new_with_thickness = {}

    patterns = [
        r"(\d{1,2}) (\b[a-zA-Z0-9\-\|_]+\b)\s+(.*)     (.{1,9}) (\d{1,5}) (\d{1,5}\.\d{1,5})",  # 0
        r"(\d{1,2}) (\b[a-zA-Z0-9\-\|_]+\b)\s+(.*)     (.{1,9}) (\d{1,5}\.\d{1,5}) (\d{1,5})",  # 1
        r"^(\d+) (\d+) (\S+) (.*?) (\d+(?:\.\d+)?)?$",  # 2
        r"^(\d+) (\S+) (\d+\.\d+) (.*?) (\d+(?:\.\d+)?)?$",  # 3
    ]
    text = convert_pdf_to_text(pdf_path=path_to_pdf_file)

    matches = []
    for i, pattern in enumerate(patterns):
        if i == 0:
            matches = re.findall(pattern, text)
            if len(matches) < 2:
                continue
            for match in matches:
                item_no, part_name, material, thickness, quantity, weight = match
                part_names_new_with_thickness[part_name] = {
                    "thickness": _make_name_safe(str(thickness)),
                    "quantity": quantity,
                    "material": material,
                }
        elif i == 1:
            matches = re.findall(pattern, text)
            if len(matches) < 2:
                continue
            for match in matches:
                item_no, part_name, material, thickness, weight, quantity = match
                part_names_new_with_thickness[part_name] = {
                    "thickness": _make_name_safe(str(thickness)),
                    "quantity": quantity,
                    "material": material,
                }
        elif i == 2:
            matches = re.findall(pattern, text, re.MULTILINE | re.DOTALL)
            if len(matches) < 2:
                continue
            for match in matches:
                if len(match) != 5 or match[4] == "" or len(match[2]) < 2:
                    continue
                item_no, quantity, part_name, material, thickness = match
                part_names_new_with_thickness[part_name] = {
                    "thickness": _make_name_safe(str(thickness)),
                    "quantity": quantity,
                    "material": material,
                }
        elif i == 3:
            for match in matches:
                if len(match) != 5 or match[4] == "" or len(match[2]) < 2:
                    continue
                item_no, quantity, part_name, material, thickness = match
                part_names_new_with_thickness[part_name] = {
                    "thickness": _make_name_safe(str(thickness)),
                    "quantity": quantity,
                }

    return part_names_new_with_thickness


def convert_pdf_to_text(pdf_path: str) -> str:
    with open("output.txt", "w") as f:
        f.write("")

    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        num_pages = len(reader.pages)

        for page_num in range(num_pages):
            page = reader.pages[page_num]
            text = page.extract_text()

            # Identify table boundaries based on specific patterns or coordinates
            # Extract table content using string manipulation or regular expressions
            # Append extracted table data to the tables list

            with open("output.txt", "a") as f:
                f.write(text)

    with open("output.txt", "r") as f:
        all_text = f.read()

    with open("output.txt", "w") as f:
        f.write(all_text)
    return all_text


def get_all_file_paths_from_directory(directory_to_sort: str) -> list[str]:
    file_paths = []
    for root, directories, files in os.walk(directory_to_sort):
        for file in files:
            file_path = os.path.join(root, file)
            file_paths.append(file_path)
    return file_paths


def filter_file_paths(file_paths: list[str], desired_file_names: list[str]) -> list[str]:
    filtered_paths = []
    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        name_without_extension = os.path.splitext(file_name)[0]
        if name_without_extension in desired_file_names:
            filtered_paths.append(file_path)
    return filtered_paths


def sort_jobs(path_to_file: str, directory_to_sort: str, output_directory: str) -> None:
    data = {}
    if path_to_file.lower().endswith(".xlsx"):
        data = get_data_from_excel(path_to_file)
    elif path_to_file.lower().endswith(".pdf"):
        data = get_data_from_pdf(path_to_file)
    all_file_paths = get_all_file_paths_from_directory(directory_to_sort=directory_to_sort)
    filtered_paths = filter_file_paths(all_file_paths, list(data.keys()))
    new_copy_location = {}
    for file_path in filtered_paths:
        file_name: str = os.path.basename(file_path)
        name_without_extension: str = os.path.splitext(file_name)[0]
        extension: str = os.path.splitext(file_name)[1].replace(".", "").upper()
        quantity_mulitplier: str = f" x{data[name_without_extension]['quantity']}" if int(data[name_without_extension]["quantity"]) > 1 else ""
        new_copy_location[file_path] = {
            "new_location": f"{output_directory}\\{extension}\\{data[name_without_extension]['thickness']}",
            "old_name": f"{output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{file_name}",
            "new_name": f"{output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{name_without_extension}{quantity_mulitplier}{os.path.splitext(file_name)[1]}",
        }

    # copy_files(new_copy_location)


def copy_files(locations_dictionary: dict) -> None:
    for original_location in locations_dictionary:
        new_location = locations_dictionary.get(original_location)["new_location"]
        old_name = locations_dictionary.get(original_location)["old_name"]
        new_name = locations_dictionary.get(original_location)["new_name"]
        os.makedirs(new_location, exist_ok=True)
        shutil.copy(original_location, new_location)
        if "DXF" in new_location:
            try:
                os.rename(old_name, new_name)
            except FileExistsError:
                os.remove(old_name)


if __name__ == "__main__":
    sort_jobs(
        path_to_file=r"C:\Users\jared\Downloads\Assy_SD10100.pdf",
        directory_to_sort=r"F:\Code\Python-Projects\Inventory Manager\test\OCFAB\Dumpsters",
        output_directory=r"C:\Users\jared\Downloads\I am Job Name",
    )
    print("done")
