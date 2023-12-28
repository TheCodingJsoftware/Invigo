import os
import re
import shutil
import traceback
from itertools import islice

from openpyxl import load_workbook
from PIL import Image
from PyQt6.QtCore import QThread, pyqtSignal


class JobSorterThread(QThread):
    """
    Uploads client data to the server
    """

    signal = pyqtSignal(object)

    def __init__(self, parent, job_name: str, path_to_excel_file: str, directory_to_sort: str, output_directory: str) -> None:
        """
        The function is a constructor for a class that inherits from QThread. It takes a list of strings
        as an argument and returns None

        Args:
          file_to_upload (list[str]): list[str] = list of files to upload
        """
        QThread.__init__(self)
        self.data = {}
        self.path_to_excel_file: str = path_to_excel_file
        self.directory_to_sort: str = directory_to_sort
        self.job_name: str = job_name
        self.output_directory: str = os.path.join(output_directory, self.job_name)

    def _make_name_safe(self, name: str):
        """
        The function replaces any invalid characters in a given string with underscores to make it safe for
        use as a directory name.

        Args:
        name (str): A string representing a directory name that needs to be made safe for use in a file
        system. The function replaces any invalid characters in the name with an underscore character.

        Returns:
        a string that is a safe version of the input `name` string, where any invalid characters (such as
        `<`, `>`, `:`, `"`, `/`, `\`, `|`, `?`, `*`) have been replaced with underscores (`_`).
        """
        invalid_chars_pattern = r'[<>:"/\\|?*]'

        safe_directory_name = re.sub(invalid_chars_pattern, "_", name)

        return safe_directory_name

    def _find_column(self, path_to_excel_file: str, column_name_to_find: str) -> int:
        """
        This function finds the column number of a given column name in an Excel file.

        Args:
        path_to_excel_file (str): A string representing the file path to the Excel file that needs to be
        searched for the column.
        column_name_to_find (str): The name of the column that needs to be found in the Excel file.

        Returns:
        an integer representing the column number where the column_name_to_find is found in the first row
        of the worksheet. If the column_name_to_find is not found in the first row, the function returns
        None.
        """
        workbook = load_workbook(path_to_excel_file)
        worksheet = workbook.active

        try:
            for column, cell in enumerate(worksheet[1]):
                if column_name_to_find.lower() in cell.value.lower():
                    return column
        except TypeError:
            return -1
        return -1

    def get_data_from_excel(self, path_to_excel_file: str) -> dict:
        """
        This function reads data from an Excel file and returns a dictionary of part names with their
        corresponding thickness.

        Args:
        path_to_excel_file (str): The path to the Excel file that contains the data to be extracted.

        Returns:
        a dictionary where the keys are part names and the values are their corresponding thicknesses,
        obtained from an Excel file specified by the input parameter `path_to_excel_file`.
        """
        workbook = load_workbook(filename=path_to_excel_file)
        worksheet = workbook.active
        part_names_new_with_thickness = {}
        part_name_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Part")
        material_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Material")
        thickness_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Thick")
        quantity_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Parts Per")
        if quantity_column == -1:
            quantity_column: int = self._find_column(path_to_excel_file=path_to_excel_file, column_name_to_find="Qty")
        for row in islice(worksheet.iter_rows(values_only=True), 1, None):
            part_name: str = row[part_name_column]
            material: str = row[material_column]
            thickness: str = row[thickness_column]
            quantity: str = row[quantity_column]

            part_names_new_with_thickness[part_name] = {"thickness": self._make_name_safe(str(thickness)), "quantity": quantity}

        return part_names_new_with_thickness

    def get_all_file_paths_from_directory(self, directory_to_sort: str) -> list[str]:
        """
        This function retrieves all file paths from a given directory and its subdirectories.

        Args:
        directory_to_sort (str): A string representing the path of the directory from which we want to
        retrieve all file paths.

        Returns:
        a list of file paths (as strings) for all files in the directory and its subdirectories specified
        by the input parameter `directory_to_sort`.
        """
        file_paths = []
        for root, directories, files in os.walk(directory_to_sort):
            for file in files:
                file_path = os.path.join(root, file)
                file_paths.append(file_path)
        return file_paths

    def filter_file_paths(self, file_paths: list[str], desired_file_names: list[str]) -> list[str]:
        """
        This function filters a list of file paths based on a list of desired file names.

        Args:
        file_paths (list[str]): A list of file paths (strings) that you want to filter.
        desired_file_names (list[str]): The list of file names that we want to filter the file paths by.

        Returns:
        a list of file paths that match the desired file names.
        """
        filtered_paths = []
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            name_without_extension = os.path.splitext(file_name)[0]
            if name_without_extension in desired_file_names:
                filtered_paths.append(file_path)
        return filtered_paths

    def run(self) -> None:
        """
        This is a Python function that extracts data from PDF files and stores it in a dictionary.
        """
        try:
            data = self.get_data_from_excel(self.path_to_excel_file)
            all_file_paths = self.get_all_file_paths_from_directory(directory_to_sort=self.directory_to_sort)
            filtered_paths = self.filter_file_paths(all_file_paths, list(data.keys()))
            new_copy_location = {}
            for file_path in filtered_paths:
                file_name: str = os.path.basename(file_path)
                name_without_extension: str = os.path.splitext(file_name)[0]
                extension: str = os.path.splitext(file_name)[1].replace(".", "").upper()
                quantity_mulitplier: str = (
                    f" x{data[name_without_extension]['quantity']}" if int(data[name_without_extension]["quantity"]) > 1 else ""
                )
                new_copy_location[file_path] = {
                    "new_location": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}",
                    "old_name": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{file_name}",
                    "new_name": f"{self.output_directory}\\{extension}\\{data[name_without_extension]['thickness']}\\{name_without_extension}{quantity_mulitplier}{os.path.splitext(file_name)[1]}",
                }

            for original_location in new_copy_location:
                new_location = new_copy_location.get(original_location)["new_location"]
                old_name = new_copy_location.get(original_location)["old_name"]
                new_name = new_copy_location.get(original_location)["new_name"]
                os.makedirs(new_location, exist_ok=True)
                shutil.copy(original_location, new_location)
                if "DXF" in new_location:
                    try:
                        os.rename(old_name, new_name)
                    except FileExistsError:
                        os.remove(old_name)

            self.signal.emit("Done")
        except Exception as e:
            self.signal.emit(f"ERROR!\nException: {e}\nTrace stack:\n{traceback.print_exc()}\n\n")
